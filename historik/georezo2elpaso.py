# -*- coding: UTF-8 -*-
#!/usr/bin/env python
from __future__ import (unicode_literals, print_function)
#------------------------------------------------------------------------------
# Name:         GeoRezo to El Paso
#
# Purpose:      Toolbox to import GeoRezo database, checking integrity and
#               generating report about conflicts between data automatically
#               processed by El Paso and data manually fixed sometimes.
#
# Authors:      Guts (https://github.com/Guts)
#
# Python:       2.7.x
# Created:      19/06/2015
# Updated:      21/06/2015
#
# Licence:      GPL 3
#------------------------------------------------------------------------------

###############################################################################
########### Libraries #############
###################################

# Standard library
from collections import Counter
from datetime import datetime
from os import path
import re
import sqlite3
import sys
import time

# 3rd party
from openpyxl import load_workbook
import pytz

###############################################################################
########## Main program ###########
###################################

## Globals
# timezone
paris_tz = pytz.timezone("Europe/Paris")


class GeorezoToElPaso(object):
    """
    source_path = fichier Excel (.xlsx) dont la structure est la suivante :

    ######### Structure attendue ############################
    # col_idx   col_name        description
    ------------------------------------------------------
    # 0         ID_def          id annonce dans la base source
    # 1         ID              id annonce dans le forum
    # 2         PID             id annonce dans le RSS
    # 3         ENTETE_OFFRE    titre de l'annonce
    # 4         DATE            date de publication
    # 5         ANNEE           année de publication
    # 6         TYPE_OFFRE      type de contrat
    # 7         VISITES         nombre de clics sur l'annonce
    # 8         DEPART_1        département identifié en 1er
    # 9         DEPART_2        département identifié en 2è
    # 10        REGION          région identifiée ou déduite
    # 11        TYPE_REGION     type de région (métro / dom)
    # 12        NBR             ?
    # 13        MESSAGE         texte intégral de l'annonce
    ########################################################

    db = base de données dans laquelle doivent être importées les données
    db_type = format du SGBD utilisé
    mode = pour indiquer le niveau d'actions à réaliser :
            \t- check-source
            \t- check-all
            \t- copy
            \t- report
            \t- import
    mirror_table = nom de la table où seront copiées les données sources
    """
    def __init__(self, source_path="toImport/BDD_JOB.xlsx",
                 db=r"../elpaso.sqlite", db_type="sqlite",
                 mirror_table="georezo_histo", mode="check-source"):
        super(GeorezoToElPaso, self).__init__()

        # testing source
        if not path.isfile(source_path)\
           and not path.splitext(source_path)[1] == ".xlsx":
            print("Source non conforme : fichier xslx attendu.")
            sys.exit()
        else:
            self.source_path = path.abspath(source_path)
            pass

        # testing mode parameter
        tup_modes = ("check-source",
                     "check-all",
                     "copy",
                     "report",
                     "import")
        if mode not in tup_modes:
            print("Le mode indiqué doit être l'un de ceux là : {0}".format(tup_modes))
            sys.exit()
        else:
            pass

        # checking source file and quit
        if mode == "check-source":
            src_report = self.check_integrity_source(self.source_path)
            if src_report[0]:
                print("{0} erreurs détectées dans le fichier source. Corriger avant d'effectuer un import.\n\n{1}".format(src_report[0], src_report[1]))
            else:
                print("")
            sys.exit()
            return
        else:
            pass

        # testing database
        if db_type == "sqlite"\
           and not path.isfile(db):
            print("Base de données non conforme : fichier SQLite attendu.")
            sys.exit()
        else:
            db = path.abspath(db)
            self.conn = sqlite3.connect(db)
            self.db_cursor = self.conn.cursor()
            pass

        # testing if dedicated table exists or not
        self.db_cursor.execute("SELECT name\
                                FROM sqlite_master\
                                WHERE type='table'\
                                ORDER BY Name")
        tables = map(lambda t: t[0], self.db_cursor.fetchall())
        if not mirror_table in tables:
            print("La table miroir indiquée ({0}) n'existe pas.".format(mirror_table))
            prompt_creation = raw_input("Ajouter ? (o/n)")
            if prompt_creation == "o":
                self.create_table(mirror_table)
            elif prompt_creation == "n":
                pass
        else:
            pass

        # closing connection
        self.conn.close()

    def create_table(self, mirror_table):
        """
        Create the dedicated table in the database if not exists already
        """
        self.db_cursor.execute("CREATE  TABLE  IF NOT EXISTS 'main'.'{0}' \
            ('idu' INTEGER PRIMARY KEY  AUTOINCREMENT  NOT NULL  UNIQUE ,\
             'id_forum' INTEGER UNIQUE ,\
             'id_rss' INTEGER UNIQUE ,\
             'title' VARCHAR,\
             'published' DATETIME,\
             'contrat' VARCHAR,\
             'visites' INTEGER,\
             'dpt1' VARCHAR,\
             'dpt2' VARCHAR,\
             'region' VARCHAR,\
             'region_typ' VARCHAR,\
             'summary' TEXT)".format(mirror_table))
        # saving changes
        self.conn.commit()
        # end of function
        return

    def reset_table(self, mirror_table):
        """
        Empty the dedicated table in the database to perform a fresh import,
        overwriting everything
        """
        # clean table
        self.db_cursor.execute("DELETE FROM '{0}'".format(mirror_table))

        # saving changes
        self.conn.commit()

        # end of function
        return

    def copy_safe(self, worksheet_path):
        """
        Just copy all data from the source workbook into a dedicated
        table in the database without touch anything to the main tables.
        """
        # opening the source file
        wb = load_workbook(filename=worksheet_path,
                           read_only=True
                           )
        ws = wb.worksheets[0]  # ws = première feuille

        # parsing line by line
        for row in ws.iter_rows(row_offset=1):
            # parcours ligne par ligne
            idu = int(row[0].value)
            try:
                id_forum = int(row[1].value)
            except:
                print('oups')
                li_issues.append(("TypeProblem", idu))
                continue
            id_rss = int(row[2].value)
            title = row[3].value
            contrat = row[6].value
            visits = row[7].value
            dpt1 = row[8].value
            dpt2 = row[9].value
            region = row[10].value
            region_typ = row[11].value

            # récupérer et nettoyer le contenu de l'offre
            summary = row[13].value
            summary_clean = self.clean_offer(summary)

            # récupérer la date de publication
            date_pub = row[4].value
            # ajouter fuseau horaire
            date_pub = paris_tz.localize(date_pub)
            published = time.strptime('{0} {1} {2}'.format(date_pub.year,
                                                           date_pub.month,
                                                           date_pub.day),
                                                           '%Y %m %d')
            if "2" in str(date_pub.utcoffset()):
                print('youhou')
                published = time.strftime("%a, %d %b %Y %H:%M:%S +0200", published)
            elif "1" in str(date_pub.utcoffset()):
                print('youpi')
                published = time.strftime("%a, %d %b %Y %H:%M:%S +0100", published)
            else:
                pass

            # insertion BDD
            try:
                self.db_cursor.execute("INSERT INTO histo_georezo\
                                        VALUES (?,?,?,?,?,?,?,?,?,?,?,?)",
                                       (str(idu),
                                        str(id_forum),
                                        str(id_rss),
                                        title,
                                        published,
                                        contrat,
                                        str(visits),
                                        dpt1,
                                        dpt2,
                                        region,
                                        region_typ,
                                        summary))
                self.conn.commit()
            except Exception as e:
                print('oups2: ', e)
                li_issues.append(("NotUnique", idu, id_rss))

        # end of method
        return

    def import_final(self):
        """
        ff
        """
        pass

    def check_integrity_source(self, worksheet_path):
        """
        ensure that the source file is healthy
        """
        li_issues = []
        src_report = {'erreurs': li_issues}

        li_id_u = []
        li_id_f = []
        li_id_r = []

        # 0         ID_def          id annonce dans la base source
        # 1         ID              id annonce dans le forum
        # 2         PID             id annonce dans le RSS
        # 3         ENTETE_OFFRE    titre de l'annonce
        # 4         DATE            date de publication
        # 5         ANNEE           année de publication
        # 6         TYPE_OFFRE      type de contrat
        # 7         VISITES         nombre de clics sur l'annonce
        # 8         DEPART_1        département identifié en 1er
        # 9         DEPART_2        département identifié en 2è
        # 10        REGION          région identifiée ou déduite
        # 11        TYPE_REGION     type de région (métro / dom)
        # 12        NBR             ?
        # 13        MESSAGE         texte intégral de l'annonce

        # opening the source file
        wb = load_workbook(filename=worksheet_path,
                           read_only=True
                           )
        ws = wb.worksheets[0]  # ws = première feuille

        # worksheet dimensions
        src_report['cols_count'] = ws.get_highest_column() + 1
        src_report['rows_count'] = ws.get_highest_row() - 1

        # checking if identifiers are unique
        for row in ws.iter_rows(row_offset=1):
            # IDs uniques de la BD
            try:
                li_id_u.append(int(row[0].value))
            except:
                li_issues.append((0, "ID_def absent, ligne {0}".format(row[3].row)))
                continue

            # IDs du forum
            try:
                li_id_f.append(int(row[1].value))
            except:
                li_issues.append((1, "ID Forum absent, ligne {0}".format(row[3].row)))
                continue

            # IDs du RSS
            try:
                li_id_r.append(int(row[2].value))
            except:
                li_issues.append((2, "ID RSS absent, ligne {0}".format(row[3].row)))
                continue

        # tuplisation to perform set more quickly
        li_id_u = tuple(li_id_u)
        li_id_f = tuple(li_id_f)
        li_id_r = tuple(li_id_r)

        # # checking if IDs are unique or not
        not_uniq_id_u = [item for item, count in Counter(li_id_u).items() if count > 1]
        if len(not_uniq_id_u) > 0:
            li_issues.append((0, "Identifiants base non uniques : {0}".format(not_uniq_id_u)))
        else:
            pass

        not_uniq_id_f = [item for item, count in Counter(li_id_f).items() if count > 1]
        if len(not_uniq_id_f) > 0:
            li_issues.append((1, "Identifiants forum non uniques : {0}".format(not_uniq_id_f)))
        else:
            pass

        not_uniq_id_r = [item for item, count in Counter(li_id_r).items() if count > 1]
        if len(not_uniq_id_r) > 0:
            li_issues.append((2, "Identifiants RSS non uniques : {0}".format(not_uniq_id_r)))
        else:
            pass

        # end of method
        return len(li_issues), src_report

    def check_report_comparison(self):
        """
        ff
        """
        pass

    def clean_offer(self, offer_text):
        """
        """
        # enlever les astérisques
        rx = re.compile('\*+')
        offer_text_clean = rx.sub(' ', offer_text).strip()

        # remplacer les doubles guillemets
        rx = re.compile('\Â«|Â»')
        offer_text_clean = rx.sub(u'"', offer_text_clean).strip()

        # enlever les copyright
        rx = re.compile('\Â®')
        offer_text_clean = rx.sub('', offer_text_clean).strip()

        # enlever les puces mal interprétées
        rx = re.compile('\â€¢')
        offer_text_clean = rx.sub('', offer_text_clean).strip()
        rx = re.compile('\â€”')
        offer_text_clean = rx.sub('', offer_text_clean).strip()
        rx = re.compile('\Â·')
        offer_text_clean = rx.sub('', offer_text_clean).strip()
        rx = re.compile('\Â°+')
        offer_text_clean = rx.sub('', offer_text_clean).strip()

        # enlever les lignes construites avec des underscores multiples
        rx = re.compile('_{2,}')
        offer_text_clean = rx.sub('', offer_text_clean).strip()

        # enlever les lignes construites avec des tirets multiples
        rx = re.compile('-{2,}')
        offer_text_clean = rx.sub('', offer_text_clean).strip()

        # enlever les codes html des caractères spéciaux mal interprétés
        rx = re.compile('&#+[0-9]{5}|;')
        offer_text_clean = rx.sub(' ', offer_text_clean).strip()
        rx = re.compile('â€¦')
        offer_text_clean = rx.sub('', offer_text_clean).strip()

        # enlever les balises html url et img
        rx = re.compile('\[[a-z]{3}\]|\[/[a-z]{3}\]')
        offer_text_clean = rx.sub(' ', offer_text_clean).strip()

        # remplacer les â
        rx = re.compile('\à¢+')
        offer_text_clean = rx.sub(u'â', offer_text_clean).strip()

        # remplacer les À
        rx = re.compile('\à€+')
        offer_text_clean = rx.sub(u'À', offer_text_clean).strip()

        # remplacer les ë
        rx = re.compile('\à«+')
        offer_text_clean = rx.sub(u'ë', offer_text_clean).strip()

        # remplacer les ê
        rx = re.compile('àª')
        offer_text_clean = rx.sub(u'ê', offer_text_clean).strip()

        # remplacer les î
        rx = re.compile('\à®+')
        offer_text_clean = rx.sub(u'î', offer_text_clean).strip()

        # remplacer les ï
        rx = re.compile('\à¯')
        offer_text_clean = rx.sub(u'ï', offer_text_clean).strip()

        # remplacer les ô
        rx = re.compile('\à´+')
        offer_text_clean = rx.sub(u'ô', offer_text_clean).strip()

        # remplacer les œ
        rx = re.compile('\Å“')
        offer_text_clean = rx.sub(u'œ', offer_text_clean).strip()

        # remplacer les û
        rx = re.compile('\à»+')
        offer_text_clean = rx.sub(u'û', offer_text_clean).strip()

        # remplacer les ù
        rx = re.compile('\à¹')
        offer_text_clean = rx.sub(u'ù', offer_text_clean).strip()

        # remplacer les €
        rx = re.compile('\â‚¬+')
        offer_text_clean = rx.sub(u' euros', offer_text_clean).strip()

        # remplacer les ²
        rx = re.compile('\Â²+')
        offer_text_clean = rx.sub(u'²', offer_text_clean).strip()

        # remplacer les apostrophes mal interprétés
        rx = re.compile('d\?i|d\?I')
        offer_text_clean = rx.sub(u"d'i", offer_text_clean).strip()
        rx = re.compile('à¢â‚¬â„')
        offer_text_clean = rx.sub(u"'", offer_text_clean).strip()

        # end of method
        return offer_text_clean

###############################################################################
###### Stand alone program ########
###################################

if __name__ == '__main__':
    """
    To launch this script from the terminal
    """
    GeorezoToElPaso(source_path="toImport/BDD_JOB.xlsx",
                    db=r"elpaso_toWork.sqlite",
                    db_type="sqlite",
                    mirror_table="georezo_histo",
                    mode="check-source")
